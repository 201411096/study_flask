#pip install openpyxl
from openpyxl import Workbook
import calendar
import datetime
import random

wb = Workbook()
ws = wb.active

columnList = ['근무일자', '사원번호', '성명', '구분', '근무형태', '근무유형', '출근시각', '퇴근시각', '작일퇴근', '수정출근시간', '수정퇴근시간', '출근IP', '퇴근IP']

# A ~ Z : 65 - 91
# a ~ z : 97 - 123
startRow = 2
startCol = 66

saved_startCol = startCol
currentTime = datetime.datetime.now()
currentYear = currentTime.year
currentMonth = currentTime.month

empNo = 'kys123'
empName = '김용선'
empIp = '123.124.345'
# empNo = input('사원번호 입력 : ')
# empName = input('사원이름 입력 : ')
# empIp = input('출퇴근IP 입력 : ')
currentMonthRange = calendar.monthrange(int(currentYear), int(currentMonth))[1]

print('==============================================================================')
for i in range(len(columnList)):    
    ws.column_dimensions[str(chr(startCol))].width = 15
    print('cellName : ',  str(chr(startCol))+str(startRow) )
    ws[ str(chr(startCol))+str(startRow) ] = columnList[i]
    startCol+=1

startCol = saved_startCol
print('==============================================================================')
for i in range(currentMonthRange):
    startCol = saved_startCol
    startRow = startRow + 1
    for j in range(len(columnList)):
        print('cellName : ',  str(chr(startCol))+str(startRow) )
        if j == 0:
            ws[ str(chr(startCol))+str(startRow) ] = str(datetime.date(currentYear, currentMonth, i+1))
        if j == 1:
            ws[ str(chr(startCol))+str(startRow) ] = empNo
        if j == 2:
            ws[ str(chr(startCol))+str(startRow) ] = empName
        if j == 6:
            hour = 8
            randomMin = random.randrange(15, 59)
            randomSec = random.randrange(0, 59)
            ws[ str(chr(startCol))+str(startRow) ] = str(datetime.time(hour, randomMin, randomSec))
        if j == 7:
            hour = 18
            randomMin = random.randrange(0, 59)
            randomSec = random.randrange(0, 59)            
            ws[ str(chr(startCol))+str(startRow) ] = str(datetime.time(hour, randomMin, randomSec))
        if j == 11:
            ws[ str(chr(startCol))+str(startRow) ] = empIp
        if j == 12:
            ws[ str(chr(startCol))+str(startRow) ] = empIp
        startCol+=1



# fileName = input('이름 입력 : ')
fileName = str(currentYear) + "년" + str(currentMonth) + "월_근태기록_" + empName
wb.save(fileName + ".xlsx")
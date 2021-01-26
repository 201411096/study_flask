#pip install openpyxl
from openpyxl import Workbook, load_workbook
import calendar
import datetime
import random
import json

def makeFile(empNo, empName, empIp, holidayList):
    # wb = Workbook()
    wb = load_workbook('./test_09_excel/template.xlsx')
    ws = wb.active

    columnList = ['근무일자', '사원번호', '성명', '구분', '근무형태', '근무유형', '출근시각', '퇴근시각', '작일퇴근', '수정출근시간', '수정퇴근시간', '출근IP', '퇴근IP']

    # A ~ Z : 65 - 91
    # a ~ z : 97 - 123
    startRow = 3
    startCol = 66

    saved_startCol = startCol
    currentTime = datetime.datetime.now()
    currentYear = currentTime.year
    currentMonth = currentTime.month
    currentMonthRange = calendar.monthrange(int(currentYear), int(currentMonth))[1]

    # 위에 컬럼 작성해주던 부분
    # for i in range(len(columnList)):    
    #     ws.column_dimensions[str(chr(startCol))].width = 15
    #     print('cellName : ',  str(chr(startCol))+str(startRow) )
    #     ws[ str(chr(startCol))+str(startRow) ] = columnList[i]
    #     startCol+=1

    startCol = saved_startCol

    for i in range(currentMonthRange):
        for j in range(len(columnList)):
            print('cellName : ',  str(chr(startCol))+str(startRow) )
            currentWeekDay = datetime.date(currentYear, currentMonth, i+1).weekday()
            if currentWeekDay == 5 or currentWeekDay == 6:
                startRow= startRow - 1
                break
            if (i+1) in holidayList:
                startRow= startRow - 1
                break
            if j == 0:
                ws[ str(chr(startCol))+str(startRow) ] = str(datetime.date(currentYear, currentMonth, i+1))
            if j == 1:
                ws[ str(chr(startCol))+str(startRow) ] = empNo
            if j == 2:
                ws[ str(chr(startCol))+str(startRow) ] = empName
            if j == 6:
                hour = 8
                randomMin = random.randrange(15, 59)
                randomSec = random.randrange(0, 59)
                ws[ str(chr(startCol))+str(startRow) ] = str(datetime.time(hour, randomMin, randomSec))
            if j == 7:
                hour = 18
                randomMin = random.randrange(0, 59)
                randomSec = random.randrange(0, 59)            
                ws[ str(chr(startCol))+str(startRow) ] = str(datetime.time(hour, randomMin, randomSec))
            if j == 11:
                ws[ str(chr(startCol))+str(startRow) ] = empIp
            if j == 12:
                ws[ str(chr(startCol))+str(startRow) ] = empIp
            startCol+=1
        startCol = saved_startCol
        startRow = startRow + 1

    fileName = str(currentYear) + "년" + str(currentMonth) + "월_근태기록_" + empName
    wb.save(fileName + ".xlsx")

with open('./test_09_excel/config.json', 'r', encoding="utf-8" ) as f:
    jsonData = json.load(f)

configData = jsonData['data']

for data in configData:
    print(data['empName'] +'의 근태일지 작성중 ...' )
    totalHoliday = input('총 휴가일수(+연휴일) 입력: ')
    holidayList = []
    for i in range(int(totalHoliday)):
        holiday = input('휴가일 입력 :')
        holidayList.append(int(holiday))
    makeFile(data['empNo'], data['empName'], data['empIp'], holidayList)



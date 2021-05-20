#pip install openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Color, Font
import calendar
import datetime
import random
import json

def makeFile(empNo, empName, empIp, holidayList, **kwargs):
    # wb = Workbook()
    # wb = load_workbook('template.xlsx')
    wb = load_workbook('./ex/ex_29_excel-commuteManagement/210126_excel_commuteManagement/test_09_excel/template.xlsx')
    ws = wb.active

    columnList = ['근무일자', '사원번호', '성명', '구분', '근무형태', '근무유형', '출근시각', '퇴근시각', '작일퇴근', '수정출근시간', '수정퇴근시간', '출근IP', '퇴근IP']

    # A ~ Z : 65 - 91
    # a ~ z : 97 - 123
    startRow = 3
    startCol = 66

    saved_startCol = startCol
    currentTime = datetime.datetime.now()
    currentYear = currentTime.year
    currentMonth = currentTime.month
    currentMonthRange = calendar.monthrange(int(currentYear), int(currentMonth))[1]

    if kwargs.get('year', None) is not None:
        currentYear = kwargs.get('year', None)
        if kwargs.get('month', None) is not None:
            currentMonth = kwargs.get('month', None)
            currentMonthRange = calendar.monthrange(int(currentYear), int(currentMonth))[1]

    # 위에 컬럼 작성해주던 부분
    # for i in range(len(columnList)):    
    #     ws.column_dimensions[str(chr(startCol))].width = 15
    #     print('cellName : ',  str(chr(startCol))+str(startRow) )
    #     ws[ str(chr(startCol))+str(startRow) ] = columnList[i]
    #     startCol+=1
    
    # 기본
    ws['B1'] = str(currentMonth) + "월 근태 확인서"

    tempName = ''
    for na in empName:
        tempName = tempName +" " + na

    ws['C1'] = tempName[1:]

    startCol = saved_startCol

    for i in range(currentMonthRange):
        for j in range(len(columnList)):
            # print('cellName : ',  str(chr(startCol))+str(startRow) )
            currentWeekDay = datetime.date(currentYear, currentMonth, i+1).weekday()
            if alldayFlag == '0':
                if (currentWeekDay == 5 or currentWeekDay == 6):
                    startRow= startRow - 1
                    break
                if (i+1) in holidayList:
                    startRow= startRow - 1
                    break
            if alldayFlag == '1':
                if (currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList):
                    ws[ str(chr(startCol))+str(startRow) ].fill = PatternFill(patternType='solid', fgColor=Color(redDayColor))
                    # ws[ str(chr(startCol))+str(startRow) ].font = Font(color=Color(redDayFontColor))
                    ws[ str(chr(startCol))+str(startRow) ].font = Font(color=Color(redDayFontColor), size=10)
            if j == 0:
                ws[ str(chr(startCol))+str(startRow) ] = str(datetime.date(currentYear, currentMonth, i+1))
            # if j == 1:
            #     ws[ str(chr(startCol))+str(startRow) ] = empNo
            # if j == 2:
            #     ws[ str(chr(startCol))+str(startRow) ] = empName
            if (j == 1) and not ((currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList)):
                ws[ str(chr(startCol))+str(startRow) ] = empNo
            if (j == 2)  and not ((currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList)):
                ws[ str(chr(startCol))+str(startRow) ] = empName
            if (j == 6) and not ((currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList)):
                hour = 8
                randomMin = random.randrange(30, 59)
                randomSec = random.randrange(0, 59)
                ws[ str(chr(startCol))+str(startRow) ] = str(datetime.time(hour, randomMin, randomSec))
            if (j == 7) and not ((currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList)):
                hour = 18
                randomMin = random.randrange(0, 29)
                randomSec = random.randrange(0, 59)            
                ws[ str(chr(startCol))+str(startRow) ] = str(datetime.time(hour, randomMin, randomSec))
            if (j == 11) and not ((currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList)):
                ws[ str(chr(startCol))+str(startRow) ] = empIp
            if (j == 12) and not ((currentWeekDay == 5 or currentWeekDay == 6) or ((i+1) in holidayList)):
                ws[ str(chr(startCol))+str(startRow) ] = empIp
            startCol+=1
        startCol = saved_startCol
        startRow = startRow + 1

    fileName = str(currentYear) + "년" + str(currentMonth) + "월_근태기록_" + empName
    wb.save(fileName + ".xlsx")

with open('./ex/ex_29_excel-commuteManagement/210126_excel_commuteManagement/test_09_excel/config.json', 'r', encoding="utf-8" ) as f:
# with open('./test_09_excel/config.json', 'r', encoding="utf-8" ) as f:
    jsonData = json.load(f)

configData = jsonData['data']
# holidayFlag = input('휴가체크(y/n) : ').lower()
holidayFlag = 'y' # 휴가 y / 무시 n
alldayFlag = '1' # include 1 / pass 0


redDayColor = 'FFFFFF'        # 빨간날 표시 색깔
redDayFontColor = 'FF7C80'    # 빨간날 글씨 색깔
# redDayColor = 'E8828D'        # 빨간날 표시 색깔
# redDayFontColor = 'FFFFFF'    # 빨간날 글씨 색깔

# anotherMonthFlag = input('다른달체크(y/n) : ').lower()
anotherMonthFlag = 'y'

if anotherMonthFlag == 'y':
    optionYear = int(input('년도 입력 : '))
    optionMonth = int(input('월 입력 : '))

commonHolidayList = []
commonHolidayNumber = input('연휴 일수 입력 :')

for i in range(int(commonHolidayNumber)):
    commonHoliday = input('연휴일 입력 : ')
    commonHolidayList.append(int(commonHoliday))

for data in configData:
    holidayList = []
    holidayList = commonHolidayList.copy()
    # holidayList = commonHolidayList[:]
    if holidayFlag == 'y':
        print(data['empName'] +'의 근태일지 작성중 ...' )
        totalHoliday = input('총 휴가일수(+연휴일) 입력: ')
        for i in range(int(totalHoliday)):
            holiday = input('휴가일 입력 :')
            holidayList.append(int(holiday))
    elif holidayFlag == 'n':
        pass
    if anotherMonthFlag == 'n':
        makeFile(data['empNo'], data['empName'], data['empIp'], holidayList, alldayFlag=alldayFlag, redDayColor=redDayColor, redDayFontColor=redDayFontColor)
    elif anotherMonthFlag == 'y':
        makeFile(data['empNo'], data['empName'], data['empIp'], holidayList, year=optionYear, month=optionMonth, alldayFlag=alldayFlag, redDayColor=redDayColor, redDayFontColor=redDayFontColor)


